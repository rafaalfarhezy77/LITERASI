import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

# --- 1. Load JSON ---
json_file = "RAFA ALFA RHEZY_V3925034_LITERASI3.json"
with open(json_file, "r", encoding="utf-8") as f:
    data = json.load(f)

# --- 2. Convert ke Excel ---
excel_file = "RAFA_ALFA_RHEZY_V3925034_LITERASI3_EXCELL.xlsx"
with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
    for sheet_name, value in data.items():
        pd.DataFrame(value).to_excel(writer, sheet_name=sheet_name, index=False)

# --- 3. Rapikan + Format ---
wb = load_workbook(excel_file)
for sheet in wb.sheetnames:
    ws = wb[sheet]

    # Format header (bold + fill warna abu muda)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Atur lebar kolom otomatis + margin
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 3  # kasih margin
        ws.column_dimensions[col_letter].width = adjusted_width

# --- 4. Save sebagai versi rapih ---
wb.save("RAFA_ALFA_RHEZY_V3925034_LITERASI3_RAPIH.xlsx")
